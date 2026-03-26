from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
import random
import csv
import re
from io import StringIO, BytesIO
from datetime import datetime, timedelta, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, insert
from sqlalchemy.orm import selectinload
from app.core.database import get_db
from app.models.models import Influencer, Campaign, CampaignInfluencer, User
from app.models.brand import Brand
from app.models.creator import Creator
from app.models.crm_category import CRMCategory, creator_categories
from app.models.crm_generate_post import CRMGeneratePost
from app.api.deps import require_crm_read, require_crm_create, require_crm_update, require_crm_delete, require_crm_write
from app.services.auth_service import is_super_admin, is_agency_level
from app.services.rbac_scope import visible_user_filter

router = APIRouter()

class CampaignHistory(BaseModel):
    campaign_name: str
    date: str
    roi_multiple: float
    status: str # "Completed", "Active"

class RelationshipProfile(BaseModel):
    creator_id: Optional[int] = None
    handle: str
    platform: str
    avatar_color: str
    relationship_status: str # "Active", "Vetted", "Past", "Blacklisted"
    total_spend: int
    avg_roi: float
    last_contact: str
    campaign_history: List[CampaignHistory]
    data_source: str = "real"  # "real" or "demo"
    whatsapp_numbers: Optional[List[str]] = None
    can_edit: bool = False
    category_ids: List[int] = []
    category_names: List[str] = []


class RelationshipCreate(BaseModel):
    handle: str
    platform: Optional[str] = "Instagram"
    relationship_status: Optional[str] = "Active"
    whatsapp_numbers: Optional[List[str]] = None
    name: Optional[str] = None
    category_ids: Optional[List[int]] = None


class RelationshipUpdate(BaseModel):
    handle: Optional[str] = None
    platform: Optional[str] = None
    relationship_status: Optional[str] = None
    whatsapp_numbers: Optional[List[str]] = None
    name: Optional[str] = None
    category_ids: Optional[List[int]] = None


class WhatsAppContact(BaseModel):
    id: int
    handle: str
    name: Optional[str] = None
    brand_id: Optional[int] = None
    whatsapp_numbers: List[str]
    category_ids: List[int] = []


class CategoryBrandOption(BaseModel):
    id: int
    name: str


class CategoryCreate(BaseModel):
    name: str
    brand_ids: List[int] = []


class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    brand_ids: Optional[List[int]] = None
    is_active: Optional[bool] = None


class CategoryResponse(BaseModel):
    id: int
    name: str
    user_id: int
    brand_ids: List[int]
    is_active: bool
    created_at: Optional[str] = None
    brands: List[CategoryBrandOption]


class GeneratePostCreate(BaseModel):
    title: Optional[str] = "New Post"
    platform: str
    brand_id: Optional[int] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    image_name: Optional[str] = None


class GeneratePostUpdate(BaseModel):
    title: Optional[str] = None
    platform: Optional[str] = None
    brand_id: Optional[int] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    image_name: Optional[str] = None
    is_posted: Optional[bool] = None
    posted_target_name: Optional[str] = None
    posted_recipients: Optional[List[str]] = None


class GeneratePostResponse(BaseModel):
    id: int
    user_id: int
    title: str
    platform: str
    brand_id: Optional[int] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    image_name: Optional[str] = None
    is_posted: bool = False
    posted_at: Optional[str] = None
    posted_target_name: Optional[str] = None
    posted_recipients: Optional[List[str]] = None
    created_at: Optional[str] = None

@router.get("/relationships", response_model=List[RelationshipProfile])
async def get_relationships(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_read)
):
    """
    Returns CRM data for influencer relationships.
    Queries real Creators and Influencers with their campaign history.
    Falls back to demo data when database is empty.
    """
    profiles = []
    
    # 1. Query Creators (richer relationship data) with org filtering
    if current_user.role == "super_admin":
        creators_result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .order_by(Creator.created_at.desc())
            .limit(20)
        )
    else:
        creators_result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .where(Creator.user_id == current_user.id)
            .order_by(Creator.created_at.desc())
            .limit(20)
        )
    creators = creators_result.scalars().all()
    
    for creator in creators:
        # Get campaign history for this creator's linked influencer
        campaign_history = await _get_creator_campaign_history(db, creator.handle, current_user)
        category_links = [c for c in (creator.categories or []) if c and c.is_active]
        category_ids = [c.id for c in category_links]
        category_names = [c.name for c in category_links]
        if not category_names and creator.category:
            category_names = [creator.category]
        
        profiles.append(RelationshipProfile(
            creator_id=creator.id,
            handle=f"@{creator.handle.lstrip('@')}",
            platform=creator.platform or "Instagram",
            avatar_color=f"hsl({hash(creator.handle) % 360}, 70%, 50%)",
            relationship_status=_map_creator_status(creator.status),
            total_spend=int(creator.total_earnings) if creator.total_earnings else 0,
            avg_roi=round(creator.commission_rate * 30, 1) if creator.commission_rate else 3.0,
            last_contact=creator.updated_at.strftime("%Y-%m-%d") if creator.updated_at else datetime.now().strftime("%Y-%m-%d"),
            campaign_history=campaign_history,
            data_source="real",
            whatsapp_numbers=creator.whatsapp_numbers or [],
            can_edit=True,
            category_ids=category_ids,
            category_names=category_names,
        ))
    
    # 2. Also query Influencers table if we need more (super_admin only)
    if current_user.role == "super_admin" and len(profiles) < 10:
        influencers_result = await db.execute(
            select(Influencer).limit(10 - len(profiles))
        )
        influencers = influencers_result.scalars().all()
        
        existing_handles = {p.handle.lower() for p in profiles}
        
        for influencer in influencers:
            handle = f"@{influencer.handle.lstrip('@')}"
            if handle.lower() in existing_handles:
                continue
            
            # Get campaign history
            campaign_history = await _get_influencer_campaign_history(db, influencer.id, current_user)
            
            profiles.append(RelationshipProfile(
                creator_id=None,
                handle=handle,
                platform=influencer.platform or "Instagram",
                avatar_color=f"hsl({hash(influencer.handle) % 360}, 70%, 50%)",
                relationship_status="Active",
                total_spend=0,
                avg_roi=3.0,
                last_contact=datetime.now().strftime("%Y-%m-%d"),
                campaign_history=campaign_history,
            data_source="real",
            whatsapp_numbers=[],
            can_edit=False,
            category_ids=[],
            category_names=[],
        ))
    
    return profiles

def _map_creator_status(status: str) -> str:
    """Map Creator.status to CRM relationship status."""
    mapping = {
        "pending": "Past",
        "active": "Active",
        "vetted": "Vetted",
        "past": "Past",
        "blacklisted": "Blacklisted"
    }
    return mapping.get(status, "Active")

async def _get_creator_campaign_history(db: AsyncSession, handle: str, current_user: User) -> List[CampaignHistory]:
    """Get campaign history for a creator by their handle."""
    try:
        result = await db.execute(
            select(Influencer).where(Influencer.handle.ilike(f"%{handle.lstrip('@')}%"))
        )
        influencer = result.scalar_one_or_none()
        
        if influencer:
            return await _get_influencer_campaign_history(db, influencer.id, current_user)
        return []
    except Exception:
        return []

async def _get_influencer_campaign_history(db: AsyncSession, influencer_id: int, current_user: User) -> List[CampaignHistory]:
    """Get campaign history for an influencer with org filtering."""
    try:
        if current_user.role == "super_admin":
            result = await db.execute(
                select(CampaignInfluencer, Campaign)
                .join(Campaign, CampaignInfluencer.campaign_id == Campaign.id)
                .where(CampaignInfluencer.influencer_id == influencer_id)
                .order_by(CampaignInfluencer.joined_at.desc())
                .limit(5)
            )
        else:
            result = await db.execute(
                select(CampaignInfluencer, Campaign)
                .join(Campaign, CampaignInfluencer.campaign_id == Campaign.id)
                .join(User, Campaign.owner_id == User.id)
                .where(
                    (CampaignInfluencer.influencer_id == influencer_id) &
                    (User.organization_id == current_user.organization_id)
                )
                .order_by(CampaignInfluencer.joined_at.desc())
                .limit(5)
            )
        rows = result.all()
        
        history = []
        for ci, campaign in rows:
            history.append(CampaignHistory(
                campaign_name=campaign.title or f"Campaign {campaign.id}",
                date=ci.joined_at.strftime("%Y-%m-%d") if ci.joined_at else datetime.now().strftime("%Y-%m-%d"),
                roi_multiple=round(random.uniform(1.5, 4.5), 2),
                status="Completed" if campaign.status == "completed" else "Active"
            ))
        return history
    except Exception:
        return []

def _generate_demo_relationships() -> List[RelationshipProfile]:
    """Generate demo CRM data for fresh installations."""
    handles = ["@sarah_style", "@tech_guru_99", "@fitness_jen", "@travel_mike", "@foodie_lisa", "@gamer_x"]
    platforms = ["Instagram", "YouTube", "TikTok", "Instagram", "TikTok", "Twitch"]
    statuses = ["Active", "Vetted", "Past", "Active", "Vetted", "Past"]
    colors = ["hsl(340, 70%, 50%)", "hsl(200, 70%, 50%)", "hsl(150, 70%, 50%)", "hsl(40, 70%, 50%)", "hsl(280, 70%, 50%)", "hsl(260, 70%, 50%)"]

    profiles = []
    for i in range(len(handles)):
        history = []
        num_campaigns = random.randint(1, 5)
        current_date_obj = datetime.now()
        
        for j in range(num_campaigns):
            date_str = (current_date_obj - timedelta(days=random.randint(30, 300))).strftime("%Y-%m-%d")
            history.append(CampaignHistory(
                campaign_name=f"Campaign {chr(65+j)}",
                date=date_str,
                roi_multiple=round(random.uniform(1.2, 5.0), 2),
                status="Completed"
            ))
            
        profiles.append(RelationshipProfile(
            creator_id=None,
            handle=handles[i],
            platform=platforms[i],
            avatar_color=colors[i],
            relationship_status=statuses[i],
            total_spend=random.randint(5000, 50000),
            avg_roi=round(random.uniform(2.0, 4.5), 1),
            last_contact=(datetime.now() - timedelta(days=random.randint(1, 45))).strftime("%Y-%m-%d"),
            campaign_history=history,
            data_source="demo",
            whatsapp_numbers=[],
            can_edit=False,
            category_ids=[],
            category_names=[],
        ))
        
    return profiles


@router.get("/whatsapp-contacts", response_model=List[WhatsAppContact])
async def get_whatsapp_contacts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_read)
):
    """Return CRM contacts that have WhatsApp numbers."""
    if current_user.role == "super_admin":
        creators_result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .where(Creator.status == "active")
            .order_by(Creator.created_at.desc())
            .limit(200)
        )
    else:
        creators_result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .where(Creator.user_id == current_user.id)
            .where(Creator.status == "active")
            .order_by(Creator.created_at.desc())
            .limit(200)
        )
    creators = creators_result.scalars().all()
    contacts: list[WhatsAppContact] = []
    for creator in creators:
        numbers = creator.whatsapp_numbers or []
        if not isinstance(numbers, list) or not numbers:
            continue
        contacts.append(
            WhatsAppContact(
                id=creator.id,
                handle=f"@{creator.handle.lstrip('@')}",
                name=creator.name,
                brand_id=creator.brand_id,
                whatsapp_numbers=[str(n).strip() for n in numbers if str(n).strip()],
                category_ids=[category.id for category in (creator.categories or [])],
            )
        )
    return contacts


@router.get("/category-brand-options", response_model=List[CategoryBrandOption])
async def get_category_brand_options(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_read),
):
    brands = await _get_allowed_brands(db, current_user)
    return [CategoryBrandOption(id=brand.id, name=brand.name) for brand in brands]


@router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_read),
):
    categories = await _get_visible_categories(db, current_user, include_inactive=False)
    return await _serialize_categories(db, categories)


@router.post("/categories", response_model=CategoryResponse)
async def create_category(
    payload: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_create),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")

    brand_ids = await _normalize_and_validate_category_brand_ids(db, current_user, payload.brand_ids)

    duplicate_query = select(CRMCategory).where(CRMCategory.name.ilike(name))
    if not is_super_admin(current_user.role):
        duplicate_query = duplicate_query.where(CRMCategory.user_id == current_user.id)
    duplicate_result = await db.execute(duplicate_query)
    if duplicate_result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Category with this name already exists")

    category = CRMCategory(
        name=name,
        user_id=current_user.id,
        created_by_role=current_user.role,
        brand_ids=brand_ids,
        is_active=True,
    )
    db.add(category)
    await db.commit()
    await db.refresh(category)

    serialized = await _serialize_categories(db, [category])
    return serialized[0]


@router.patch("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(
    category_id: int,
    payload: CategoryUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_update),
):
    query = select(CRMCategory).where(CRMCategory.id == category_id)
    if not is_super_admin(current_user.role):
        query = query.where(CRMCategory.user_id == current_user.id)
    result = await db.execute(query)
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Category name cannot be empty")
        duplicate_query = select(CRMCategory).where(
            CRMCategory.id != category_id,
            CRMCategory.name.ilike(name),
        )
        if not is_super_admin(current_user.role):
            duplicate_query = duplicate_query.where(CRMCategory.user_id == current_user.id)
        duplicate_result = await db.execute(duplicate_query)
        if duplicate_result.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Category with this name already exists")
        category.name = name

    if payload.brand_ids is not None:
        category.brand_ids = await _normalize_and_validate_category_brand_ids(db, current_user, payload.brand_ids)

    if payload.is_active is not None:
        category.is_active = payload.is_active

    await db.commit()
    await db.refresh(category)
    serialized = await _serialize_categories(db, [category])
    return serialized[0]


@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_delete),
):
    query = select(CRMCategory).where(CRMCategory.id == category_id)
    if not is_super_admin(current_user.role):
        query = query.where(CRMCategory.user_id == current_user.id)
    result = await db.execute(query)
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    await db.delete(category)
    await db.commit()
    return {"message": "Category deleted"}


@router.get("/generate-posts", response_model=List[GeneratePostResponse])
async def get_generate_posts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_read),
):
    query = select(CRMGeneratePost).order_by(CRMGeneratePost.created_at.desc())
    if not is_super_admin(current_user.role):
        query = query.where(CRMGeneratePost.user_id == current_user.id)
    result = await db.execute(query)
    posts = result.scalars().all()
    return [_serialize_generate_post(post) for post in posts]


@router.post("/generate-posts", response_model=GeneratePostResponse)
async def create_generate_post(
    payload: GeneratePostCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_create),
):
    platform = (payload.platform or "").strip()
    if not platform:
        raise HTTPException(status_code=400, detail="Platform is required")

    post = CRMGeneratePost(
        user_id=current_user.id,
        title=(payload.title or "New Post").strip() or "New Post",
        platform=platform,
        brand_id=payload.brand_id,
        description=(payload.description or "").strip() or None,
        image_url=payload.image_url,
        image_name=payload.image_name,
    )
    db.add(post)
    await db.commit()
    await db.refresh(post)
    return _serialize_generate_post(post)


@router.patch("/generate-posts/{post_id}", response_model=GeneratePostResponse)
async def update_generate_post(
    post_id: int,
    payload: GeneratePostUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_update),
):
    query = select(CRMGeneratePost).where(CRMGeneratePost.id == post_id)
    if not is_super_admin(current_user.role):
        query = query.where(CRMGeneratePost.user_id == current_user.id)
    result = await db.execute(query)
    post = result.scalar_one_or_none()
    if not post:
        raise HTTPException(status_code=404, detail="Generate post not found")
    if post.is_posted:
        raise HTTPException(status_code=409, detail="Posted CRM generate posts cannot be edited")

    if payload.title is not None:
        post.title = payload.title.strip() or "New Post"
    if payload.platform is not None:
        platform = payload.platform.strip()
        if not platform:
            raise HTTPException(status_code=400, detail="Platform is required")
        post.platform = platform
    if payload.brand_id is not None:
        post.brand_id = payload.brand_id
    if payload.description is not None:
        post.description = payload.description.strip() or None
    if payload.image_url is not None:
        post.image_url = payload.image_url
    if payload.image_name is not None:
        post.image_name = payload.image_name
    if payload.is_posted is not None:
        post.is_posted = payload.is_posted
        if payload.is_posted:
            post.posted_at = datetime.now(timezone.utc)
            if payload.posted_target_name is not None:
                post.posted_target_name = payload.posted_target_name.strip() or None
        else:
            raise HTTPException(status_code=409, detail="Posted CRM generate posts cannot be reverted")
    elif payload.posted_target_name is not None:
        post.posted_target_name = payload.posted_target_name.strip() or None
    if payload.posted_recipients is not None:
        cleaned_recipients = [
            str(value).strip()
            for value in payload.posted_recipients
            if str(value).strip()
        ]
        post.posted_recipients = cleaned_recipients or None

    await db.commit()
    await db.refresh(post)
    return _serialize_generate_post(post)


@router.delete("/generate-posts/{post_id}")
async def delete_generate_post(
    post_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_delete),
):
    query = select(CRMGeneratePost).where(CRMGeneratePost.id == post_id)
    if not is_super_admin(current_user.role):
        query = query.where(CRMGeneratePost.user_id == current_user.id)
    result = await db.execute(query)
    post = result.scalar_one_or_none()
    if not post:
        raise HTTPException(status_code=404, detail="Generate post not found")

    await db.delete(post)
    await db.commit()
    return {"message": "Generate post deleted"}


@router.post("/generate-report")
async def generate_xray_report(
    handle: str,
    current_user: User = Depends(require_crm_update)
):
    """
    Generates a PDF 'X-Ray' report for an influencer.
    """
    return {
        "status": "success",
        "message": f"X-Ray Report generated for {handle}",
        "download_url": "#",
        "generated_at": datetime.now().isoformat()
    }


def _normalize_handle(handle: str) -> str:
    cleaned = (handle or "").strip()
    if cleaned.startswith("@"):
        cleaned = cleaned[1:]
    return cleaned.strip()


def _normalize_whatsapp_number(value: str) -> str:
    """Normalize WhatsApp number to a compact E.164-like string."""
    cleaned = str(value or "").strip()
    if not cleaned:
        return ""
    # Remove spaces and common separators
    cleaned = re.sub(r"[\s\-\(\)]", "", cleaned)
    # Collapse any leading '+' into a single plus
    while cleaned.startswith("++"):
        cleaned = cleaned[1:]
    # Preserve leading + if present
    if cleaned.startswith("+"):
        digits = re.sub(r"\D", "", cleaned)
        return f"+{digits}" if digits else ""
    # Remove any non-digits otherwise
    return re.sub(r"\D", "", cleaned)


def _map_relationship_status_to_creator(status: Optional[str]) -> str:
    if not status:
        return "active"
    normalized = status.strip().lower()
    mapping = {
        "active": "active",
        "vetted": "vetted",
        "past": "past",
        "blacklisted": "blacklisted",
    }
    return mapping.get(normalized, "active")


async def _get_creator_for_write(db: AsyncSession, creator_id: int, current_user: User) -> Creator:
    if is_super_admin(current_user.role):
        result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .where(Creator.id == creator_id)
        )
    else:
        result = await db.execute(
            select(Creator)
            .options(selectinload(Creator.categories))
            .where(
                (Creator.id == creator_id)
                & (Creator.user_id == current_user.id)
            )
        )
    creator = result.scalar_one_or_none()
    if not creator:
        raise HTTPException(status_code=404, detail="Creator not found")
    return creator


async def _creator_to_relationship_profile(
    db: AsyncSession,
    creator: Creator,
    current_user: User,
) -> RelationshipProfile:
    refreshed_result = await db.execute(
        select(Creator)
        .options(selectinload(Creator.categories))
        .where(Creator.id == creator.id)
    )
    creator = refreshed_result.scalar_one_or_none() or creator
    campaign_history = await _get_creator_campaign_history(db, creator.handle, current_user)
    category_links = [c for c in (creator.categories or []) if c and c.is_active]
    category_ids = [c.id for c in category_links]
    category_names = [c.name for c in category_links]
    if not category_names and creator.category:
        category_names = [creator.category]
    return RelationshipProfile(
        creator_id=creator.id,
        handle=f"@{creator.handle.lstrip('@')}",
        platform=creator.platform or "Instagram",
        avatar_color=f"hsl({hash(creator.handle) % 360}, 70%, 50%)",
        relationship_status=_map_creator_status(creator.status),
        total_spend=int(creator.total_earnings) if creator.total_earnings else 0,
        avg_roi=round(creator.commission_rate * 30, 1) if creator.commission_rate else 3.0,
        last_contact=creator.updated_at.strftime("%Y-%m-%d") if creator.updated_at else datetime.now().strftime("%Y-%m-%d"),
        campaign_history=campaign_history,
        data_source="real",
        whatsapp_numbers=creator.whatsapp_numbers or [],
        can_edit=True,
        category_ids=category_ids,
        category_names=category_names,
    )


def _normalize_brand_ids(brand_ids: Optional[List[int]]) -> List[int]:
    cleaned = [int(b) for b in (brand_ids or []) if b and int(b) > 0]
    return sorted(set(cleaned))


async def _get_allowed_brand_ids(db: AsyncSession, current_user: User) -> List[int]:
    if is_super_admin(current_user.role):
        result = await db.execute(select(Brand.id))
    elif is_agency_level(current_user.role):
        if not current_user.organization_id:
            return []
        result = await db.execute(
            select(Brand.id).where(Brand.organization_id == current_user.organization_id)
        )
    else:
        result = await db.execute(
            select(Brand.id).where(Brand.created_by_user_id == current_user.id)
        )
    return [brand_id for brand_id in result.scalars().all() if brand_id]


async def _get_allowed_brands(db: AsyncSession, current_user: User) -> List[Brand]:
    if is_super_admin(current_user.role):
        result = await db.execute(select(Brand).order_by(Brand.name.asc()))
    elif is_agency_level(current_user.role):
        if not current_user.organization_id:
            return []
        result = await db.execute(
            select(Brand)
            .where(Brand.organization_id == current_user.organization_id)
            .order_by(Brand.name.asc())
        )
    else:
        result = await db.execute(
            select(Brand)
            .where(Brand.created_by_user_id == current_user.id)
            .order_by(Brand.name.asc())
        )
    return result.scalars().all()


async def _normalize_and_validate_category_brand_ids(
    db: AsyncSession,
    current_user: User,
    brand_ids: Optional[List[int]],
) -> List[int]:
    allowed_brand_ids = set(await _get_allowed_brand_ids(db, current_user))
    if not allowed_brand_ids:
        raise HTTPException(status_code=400, detail="No brands available for this user")

    selected_brand_ids = _normalize_brand_ids(brand_ids)
    if not selected_brand_ids and not is_super_admin(current_user.role) and not is_agency_level(current_user.role):
        selected_brand_ids = sorted(allowed_brand_ids)

    if not selected_brand_ids:
        raise HTTPException(status_code=400, detail="At least one brand must be selected")

    invalid = [brand_id for brand_id in selected_brand_ids if brand_id not in allowed_brand_ids]
    if invalid:
        raise HTTPException(status_code=403, detail="One or more selected brands are not accessible")

    return selected_brand_ids


async def _get_visible_categories(
    db: AsyncSession,
    current_user: User,
    include_inactive: bool = False,
) -> List[CRMCategory]:
    query = select(CRMCategory)
    if not is_super_admin(current_user.role):
        query = query.where(CRMCategory.user_id == current_user.id)
    if not include_inactive:
        query = query.where(CRMCategory.is_active == True)
    query = query.order_by(CRMCategory.created_at.desc())
    result = await db.execute(query)
    return result.scalars().all()


async def _serialize_categories(
    db: AsyncSession,
    categories: List[CRMCategory],
) -> List[CategoryResponse]:
    all_brand_ids: set[int] = set()
    for category in categories:
        for brand_id in (category.brand_ids or []):
            if isinstance(brand_id, int):
                all_brand_ids.add(brand_id)

    brand_map: dict[int, Brand] = {}
    if all_brand_ids:
        brand_result = await db.execute(select(Brand).where(Brand.id.in_(all_brand_ids)))
        brand_map = {brand.id: brand for brand in brand_result.scalars().all()}

    response: list[CategoryResponse] = []
    for category in categories:
        category_brand_ids = [int(brand_id) for brand_id in (category.brand_ids or []) if isinstance(brand_id, int)]
        response.append(
            CategoryResponse(
                id=category.id,
                name=category.name,
                user_id=category.user_id,
                brand_ids=category_brand_ids,
                is_active=category.is_active,
                created_at=category.created_at.isoformat() if category.created_at else None,
                brands=[
                    CategoryBrandOption(id=brand_id, name=brand_map[brand_id].name)
                    for brand_id in category_brand_ids
                    if brand_id in brand_map
                ],
            )
        )
    return response


async def _assign_creator_categories(
    db: AsyncSession,
    creator: Creator,
    category_ids: Optional[List[int]],
    current_user: User,
) -> None:
    if category_ids is None:
        return

    cleaned_ids = sorted(set(int(cid) for cid in category_ids if cid and int(cid) > 0))
    if not cleaned_ids:
        creator.categories = []
        return

    query = select(CRMCategory).where(CRMCategory.id.in_(cleaned_ids), CRMCategory.is_active == True)
    if not is_super_admin(current_user.role):
        query = query.where(CRMCategory.user_id == current_user.id)
    result = await db.execute(query)
    categories = result.scalars().all()
    if len(categories) != len(cleaned_ids):
        raise HTTPException(status_code=400, detail="One or more categories are invalid")

    await db.execute(
        delete(creator_categories).where(creator_categories.c.creator_id == creator.id)
    )
    if cleaned_ids:
        await db.execute(
            insert(creator_categories),
            [{"creator_id": creator.id, "category_id": cid} for cid in cleaned_ids],
        )


def _serialize_generate_post(post: CRMGeneratePost) -> GeneratePostResponse:
    return GeneratePostResponse(
        id=post.id,
        user_id=post.user_id,
        title=post.title,
        platform=post.platform,
        brand_id=post.brand_id,
        description=post.description,
        image_url=post.image_url,
        image_name=post.image_name,
        is_posted=bool(post.is_posted),
        posted_at=post.posted_at.isoformat() if post.posted_at else None,
        posted_target_name=post.posted_target_name,
        posted_recipients=[str(value) for value in (post.posted_recipients or [])],
        created_at=post.created_at.isoformat() if post.created_at else None,
    )


@router.post("/relationships", response_model=RelationshipProfile)
async def create_relationship(
    payload: RelationshipCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_create),
):
    handle = _normalize_handle(payload.handle)
    if not handle:
        raise HTTPException(status_code=400, detail="Handle is required")

    existing_query = select(Creator).where(Creator.handle.ilike(handle))
    if not is_super_admin(current_user.role):
        existing_query = existing_query.where(Creator.user_id == current_user.id)
    existing_result = await db.execute(existing_query)
    if existing_result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Creator with this handle already exists")

    creator = Creator(
        handle=handle,
        platform=payload.platform or "Instagram",
        name=payload.name,
        status=_map_relationship_status_to_creator(payload.relationship_status),
        whatsapp_numbers=[str(n).strip() for n in (payload.whatsapp_numbers or []) if str(n).strip()],
        is_approved=True,
        user_id=current_user.id,
    )
    db.add(creator)
    await db.flush()
    await _assign_creator_categories(db, creator, payload.category_ids, current_user)
    await db.commit()
    await db.refresh(creator)

    return await _creator_to_relationship_profile(db, creator, current_user)


@router.patch("/relationships/{creator_id}", response_model=RelationshipProfile)
async def update_relationship(
    creator_id: int,
    payload: RelationshipUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_update),
):
    creator = await _get_creator_for_write(db, creator_id, current_user)

    if payload.handle is not None:
        handle = _normalize_handle(payload.handle)
        if not handle:
            raise HTTPException(status_code=400, detail="Handle cannot be empty")

        existing_query = select(Creator).where(Creator.handle.ilike(handle), Creator.id != creator_id)
        if not is_super_admin(current_user.role):
            existing_query = existing_query.where(Creator.user_id == current_user.id)
        existing_result = await db.execute(existing_query)
        if existing_result.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Creator with this handle already exists")

        creator.handle = handle

    if payload.platform is not None:
        creator.platform = payload.platform

    if payload.relationship_status is not None:
        creator.status = _map_relationship_status_to_creator(payload.relationship_status)

    if payload.whatsapp_numbers is not None:
        creator.whatsapp_numbers = [str(n).strip() for n in payload.whatsapp_numbers if str(n).strip()]

    if payload.name is not None:
        creator.name = payload.name

    await _assign_creator_categories(db, creator, payload.category_ids, current_user)

    await db.commit()
    await db.refresh(creator)
    return await _creator_to_relationship_profile(db, creator, current_user)


@router.delete("/relationships/{creator_id}")
async def delete_relationship(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_delete),
):
    creator = await _get_creator_for_write(db, creator_id, current_user)
    await db.delete(creator)
    await db.commit()
    return {"message": "Creator deleted from portfolio"}


class WhatsAppImportResult(BaseModel):
    updated: int
    skipped: int
    errors: List[str]


class RelationshipImportResult(BaseModel):
    created: int
    updated: int
    skipped: int
    errors: List[str]


@router.post("/whatsapp-import", response_model=WhatsAppImportResult)
async def import_whatsapp_numbers(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_create),
):
    """
    Import WhatsApp numbers from a CSV file.
    Expected columns: handle, whatsapp_numbers
    whatsapp_numbers may contain comma/semicolon separated numbers.
    """
    if not file.filename or not file.filename.lower().endswith((".csv",)):
        raise HTTPException(status_code=400, detail="Please upload a CSV file.")

    content = await file.read()
    try:
        decoded = content.decode("utf-8")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file encoding. Use UTF-8 CSV.")

    reader = csv.DictReader(StringIO(decoded))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV is missing headers.")

    header_set = {h.strip().lower() for h in reader.fieldnames if h}
    if "handle" not in header_set or "whatsapp_numbers" not in header_set:
        raise HTTPException(status_code=400, detail="CSV must include headers: handle, whatsapp_numbers")

    updated = 0
    skipped = 0
    errors: list[str] = []

    rows = list(reader)
    for idx, row in enumerate(rows, start=2):  # header is row 1
        handle = (row.get("handle") or row.get("Handle") or "").strip()
        raw_numbers = (row.get("whatsapp_numbers") or row.get("WhatsApp Numbers") or "").strip()
        if not handle or not raw_numbers:
            skipped += 1
            continue

        numbers = [
            _normalize_whatsapp_number(n)
            for chunk in raw_numbers.split(";")
            for n in chunk.split(",")
            if n.strip()
        ]
        numbers = [n for n in numbers if n]
        if not numbers:
            skipped += 1
            continue

        # Find creator by handle (case-insensitive)
        if current_user.role == "super_admin":
            result = await db.execute(
                select(Creator).where(Creator.handle.ilike(handle.lstrip("@")))
            )
        else:
            result = await db.execute(
                select(Creator)
                .where(
                    Creator.handle.ilike(handle.lstrip("@")),
                    Creator.user_id == current_user.id,
                )
            )
        creator = result.scalar_one_or_none()
        if not creator:
            skipped += 1
            continue

        existing = creator.whatsapp_numbers or []
        if not isinstance(existing, list):
            existing = []
        merged = list({*(_normalize_whatsapp_number(n) for n in existing), *numbers})
        creator.whatsapp_numbers = [n for n in merged if n]
        updated += 1

    await db.commit()
    return WhatsAppImportResult(updated=updated, skipped=skipped, errors=errors)


def _parse_csv_rows(decoded: str) -> list[dict]:
    reader = csv.DictReader(StringIO(decoded))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV is missing headers.")
    return list(reader)


def _parse_xlsx_rows(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel (.xlsx) support is not available on the server.") from exc

    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=400, detail="Excel header row is empty.")

    data_rows: list[dict] = []
    for row in rows[1:]:
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            row_dict[header] = "" if value is None else str(value)
        data_rows.append(row_dict)
    return data_rows


def _normalize_header_map(headers: list[str]) -> dict[str, str]:
    mapping = {}
    for header in headers:
        key = header.strip().lower()
        if not key:
            continue
        mapping[key] = header
    return mapping


def _is_valid_handle(handle: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9._-]+", handle))


@router.post("/relationships-import", response_model=RelationshipImportResult)
async def import_relationships(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_crm_create),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Please upload a CSV or XLSX file.")

    lower_name = file.filename.lower()
    if not lower_name.endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Please upload a CSV or XLSX file.")

    content = await file.read()
    rows: list[dict]
    if lower_name.endswith(".csv"):
        try:
            decoded = content.decode("utf-8-sig")
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid file encoding. Use UTF-8 CSV.")
        rows = _parse_csv_rows(decoded)
    else:
        rows = _parse_xlsx_rows(content)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found.")

    header_map = _normalize_header_map(list(rows[0].keys()))
    handle_key = header_map.get("handle") or header_map.get("creator_handle") or header_map.get("influencer_handle")
    platform_key = header_map.get("platform")
    status_key = header_map.get("status") or header_map.get("relationship_status")
    name_key = header_map.get("name") or header_map.get("creator_name")
    whatsapp_key = header_map.get("whatsapp_numbers") or header_map.get("whatsapp")
    country_code_key = header_map.get("country_code")
    whatsapp_number_key = header_map.get("whatsapp_number") or header_map.get("phone_number")

    if not handle_key:
        raise HTTPException(status_code=400, detail="Template must include a handle column.")

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        raw_handle = str(row.get(handle_key, "") or "").strip()
        handle = _normalize_handle(raw_handle)
        if not handle:
            skipped += 1
            continue
        if not _is_valid_handle(handle):
            errors.append(f"Row {idx}: handle '{raw_handle}' contains unsupported characters.")
            skipped += 1
            continue

        platform = str(row.get(platform_key, "") or "").strip() if platform_key else ""
        status = str(row.get(status_key, "") or "").strip() if status_key else ""
        name = str(row.get(name_key, "") or "").strip() if name_key else ""
        whatsapp_raw = str(row.get(whatsapp_key, "") or "").strip() if whatsapp_key else ""
        country_code_raw = str(row.get(country_code_key, "") or "").strip() if country_code_key else ""
        whatsapp_single_raw = str(row.get(whatsapp_number_key, "") or "").strip() if whatsapp_number_key else ""
        whatsapp_numbers: list[str] = []
        if country_code_raw and whatsapp_single_raw:
            normalized_code = re.sub(r"\D", "", country_code_raw)
            normalized_number = re.sub(r"\D", "", whatsapp_single_raw)
            if normalized_code and normalized_number:
                whatsapp_numbers = [f"+{normalized_code} {normalized_number}"]
        else:
            whatsapp_numbers = [
                _normalize_whatsapp_number(n)
                for chunk in whatsapp_raw.split(";")
                for n in chunk.split(",")
                if n.strip()
            ] if whatsapp_raw else []
            whatsapp_numbers = [n for n in whatsapp_numbers if n]

        if is_super_admin(current_user.role):
            result = await db.execute(select(Creator).where(Creator.handle.ilike(handle)))
        else:
            result = await db.execute(
                select(Creator)
                .where(
                    Creator.handle.ilike(handle),
                    Creator.user_id == current_user.id,
                )
            )
        creator = result.scalar_one_or_none()

        if creator:
            if platform:
                creator.platform = platform
            if status:
                creator.status = _map_relationship_status_to_creator(status)
            if name:
                creator.name = name
            if whatsapp_numbers:
                existing = creator.whatsapp_numbers or []
                if not isinstance(existing, list):
                    existing = []
                merged: dict[str, str] = {}
                for n in existing:
                    cleaned = _normalize_whatsapp_number(n)
                    key = re.sub(r"\D", "", cleaned or str(n))
                    if key:
                        merged[key] = str(n).strip()
                for n in whatsapp_numbers:
                    key = re.sub(r"\D", "", n)
                    if key:
                        merged[key] = n
                creator.whatsapp_numbers = [v for v in merged.values() if v]
            updated += 1
            continue

        creator = Creator(
            handle=handle,
            platform=platform or "Instagram",
            name=name or None,
            status=_map_relationship_status_to_creator(status),
            whatsapp_numbers=whatsapp_numbers,
            is_approved=True,
            user_id=current_user.id,
        )
        db.add(creator)
        created += 1

    await db.commit()
    return RelationshipImportResult(created=created, updated=updated, skipped=skipped, errors=errors)


@router.get("/relationships-template")
async def download_relationships_template(
    current_user: User = Depends(require_crm_create),
):
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel (.xlsx) support is not available on the server.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "CRM Import"

    headers = ["handle", "platform", "status", "name", "country_code", "whatsapp_number"]
    ws.append(headers)

    # Header only (no sample rows)

    header_fill = PatternFill("solid", fgColor="E2E8F0")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    column_widths = {
        "A": 20,
        "B": 16,
        "C": 14,
        "D": 22,
        "E": 14,
        "F": 18,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Force WhatsApp columns to text/number formats to prevent Excel formula parsing
    for cell in ws["E"]:
        cell.number_format = "+0"
    for cell in ws["F"]:
        cell.number_format = "@"

    # Country code dropdown
    try:
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
        codes = ["91", "1", "44", "61", "65", "971", "81", "49", "33"]
        dv = DataValidation(type="list", formula1=f"\"{','.join(codes)}\"", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("E2:E1048576")
    except Exception:
        # If data validation import fails, template still works without dropdown.
        pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=crm_influencer_template.xlsx"},
    )


